import json
import csv
import openpyxl

def load_json():
    filename = input("Введите путь к JSON-файлу:\n")
    with open(filename, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data

def load_csv():
    filename = input("Введите путь к CSV-файлу:\n")
    with open(filename, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        return list(reader)

def load_xlsx():
    filename = input("Введите путь к XLSX-файлу:\n")
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    data = []
    for row in sheet.iter_rows(min_row=2):
        record = {}
        for header, cell in zip(headers, row):
            record[header] = cell.value
        data.append(record)
    return data